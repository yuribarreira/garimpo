from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd

from .base import BaseExtractor


class ExcelExtractor(BaseExtractor):
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        super().__init__(config)
        self.sheet_name = self.config.get("sheet_name", 0)
        self.header = self.config.get("header", 0)
        self.usecols = self.config.get("usecols")
        self.dtype = self.config.get("dtype")
        self.parse_dates = self.config.get("parse_dates")
        self.engine = self.config.get("engine")
        self.na_values = self.config.get("na_values")

    def _extract_impl(self, file_path: Union[str, Path], **kwargs):
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel não encontrado: {file_path}")

        self.logger.info(f"lendo Excel: {file_path.name}")
        params = {
            "io": file_path,
            "sheet_name": self.sheet_name,
            "header": self.header,
            "usecols": self.usecols,
            "dtype": self.dtype,
            "parse_dates": self.parse_dates,
            "engine": self.engine,
            "na_values": self.na_values,
            **kwargs,
        }
        params = {k: v for k, v in params.items() if v is not None}

        result = pd.read_excel(**params)
        if isinstance(result, dict):
            self.logger.info(f"Excel: {len(result)} planilhas")
        else:
            self.logger.info(f"Excel: {len(result)} linhas, {len(result.columns)} colunas")
        return result

    def list_sheets(self, file_path: Union[str, Path]) -> List[str]:
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel não encontrado: {file_path}")
        if file_path.suffix.lower() == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        return pd.ExcelFile(file_path, engine=self.engine).sheet_names

    def extract_all_sheets(self, file_path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
        prev = self.cache_enabled
        self.cache_enabled = False
        try:
            return self._extract_impl(file_path, sheet_name=None)
        finally:
            self.cache_enabled = prev
